import os
import zipfile
from os.path import basename
from zipfile import ZipFile
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook

path_files = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')
path_resources = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
path_zip = os.path.join(path_resources, "zip_file.zip")


def test_create_zip_archive():
    file_dir = os.listdir(path_files)
    with ZipFile(path_zip, "w") as myzip:
        for file in file_dir:
            add_file = os.path.join(path_files, file)
            myzip.write(add_file, basename(add_file))


def test_xlsx():
    with ZipFile(path_zip) as archive:
        archive.extract('xlsx_file.xlsx')
        workbook = load_workbook('xlsx_file.xlsx')
        sheet = workbook.active
        check_value = str(sheet.cell(row=1, column=1).value)
        assert check_value == 'created_at'
        os.remove('xlsx_file.xlsx')


def test_csv():
    with zipfile.ZipFile(path_zip) as archive:
        csv_archived = archive.extract('csv_file.csv')
        with open(csv_archived) as csv_file:
            csv_rows = csv.reader(csv_file)
            new_list = []
            for row in csv_rows:
                new_list.append(row)
            assert len(new_list) == 7
            assert new_list == [['Э220108-85038;453', '9'],
                               ['Э220108-43980;393', '79'],
                               ['Э220108-80594;1902', '6'],
                               ['Э220108-83110;396', '69'],
                               ['Э220108-95879;179', '9'],
                               ['Р220108-77542;249', '35'],
                               ['Э220108-30353;1207', '45']]
        os.remove('csv_file.csv')


def test_pdf():
    with ZipFile(path_zip) as archive:
        archive.extract('pdf_file.pdf')
        text = PdfReader('pdf_file.pdf').pages[0].extract_text()
        assert text.__contains__('Заявка на производство')
        os.remove('pdf_file.pdf')